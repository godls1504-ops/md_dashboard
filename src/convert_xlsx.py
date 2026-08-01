# -*- coding: utf-8 -*-
"""승인된(사전 등재) XLSX 시트를 data/converted/{시트명}.csv 로 변환.

원칙:
- 원천 XLSX 미수정 (openpyxl read_only 로딩, 저장 호출 없음)
- 값 강제 변환 없음: 셀 원본값을 그대로 기록 → ID·바코드 등은 문자열로 보존
- 원본 시트명·열 순서 유지, UTF-8(BOM, utf-8-sig), 인덱스 열 없음(index=False)
- 재실행 가능: 실행할 때마다 동일 결과로 덮어씀

"승인된 데이터 시트" = data_dictionary.table_name 에 등재된 테이블(README·data_dictionary 제외).
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

import openpyxl

# --- 경로 (추후 src/config.py 로 중앙화 가능) ---
BASE = Path(__file__).resolve().parent.parent
_CANDIDATES = [BASE / "data" / "raw" / "ganaswim_dataset.xlsx",
               BASE / "ganaswim_dataset.xlsx"]
XLSX_PATH = next((p for p in _CANDIDATES if p.exists()), _CANDIDATES[-1])
OUT_DIR = BASE / "data" / "converted"
VALIDATION_PATH = BASE / "reports" / "conversion_validation.csv"
ENCODING = "utf-8-sig"
DOC_SHEETS = {"README", "data_dictionary"}  # 데이터 테이블 아님


def cell_to_str(v) -> str:
    """셀 원본값을 강제 변환 없이 문자열화. 공란은 빈 문자열."""
    if v is None:
        return ""
    return str(v)


def load_matrix(ws) -> tuple[list, list]:
    """시트를 (헤더, 데이터행들) 로 반환. 후행 완전공백 행만 제거."""
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    while rows and all(c is None or (isinstance(c, str) and c == "") for c in rows[-1]):
        rows.pop()
    if not rows:
        return [], []
    return rows[0], rows[1:]


def approved_sheets(wb) -> list[str]:
    """data_dictionary.table_name 등재 순서대로, 실제 존재하는 시트만."""
    names: list[str] = []
    if "data_dictionary" in wb.sheetnames:
        dd = wb["data_dictionary"]
        for row in dd.iter_rows(min_row=2, values_only=True):
            t = row[0]
            if t and str(t) not in names:
                names.append(str(t))
    approved = [t for t in names if t in wb.sheetnames and t not in DOC_SHEETS]
    return approved


def _id_col_index(headers: list) -> int:
    hl = [str(h).lower() for h in headers]
    for i, h in enumerate(hl):
        if h.endswith("_id"):
            return i
    for i, h in enumerate(hl):
        if any(k in h for k in ("id", "코드", "바코드", "번호")):
            return i
    return 0


def convert_sheet(ws, out_dir: Path) -> Path:
    """한 시트를 CSV로 기록하고 경로 반환."""
    header, data = load_matrix(ws)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{ws.title}.csv"
    with open(out_path, "w", newline="", encoding=ENCODING) as f:
        w = csv.writer(f)
        w.writerow([cell_to_str(h) for h in header])
        for r in data:
            w.writerow([cell_to_str(c) for c in r])
    return out_path


def validate_sheet(ws, csv_path: Path) -> dict:
    """XLSX 원본과 기록된 CSV를 비교해 검증 결과 dict 반환."""
    x_header, x_data = load_matrix(ws)
    x_header_s = [cell_to_str(h) for h in x_header]

    with open(csv_path, "r", newline="", encoding=ENCODING) as f:
        rows = list(csv.reader(f))
    c_header = rows[0] if rows else []
    c_data = rows[1:]

    idx = _id_col_index(x_header)
    id_col = x_header_s[idx] if x_header_s else ""

    def sample_ids(header_s, data, i, n=3):
        out = []
        for r in data:
            if i < len(r) and str(r[i]).strip() != "":
                out.append(str(r[i]))
            if len(out) >= n:
                break
        return out

    x_ids = sample_ids(x_header_s, [[cell_to_str(c) for c in r] for r in x_data], idx)
    c_ids = sample_ids(c_header, c_data, idx)

    # 한글깨짐: 치환문자(U+FFFD) 없음 + 헤더 한글 일치
    raw = csv_path.read_text(encoding=ENCODING, errors="strict")
    no_replacement = "�" not in raw

    rows_match = len(x_data) == len(c_data)
    cols_match = len(x_header_s) == len(c_header)
    col_order_match = x_header_s == c_header
    id_sample_match = x_ids == c_ids
    hangul_ok = no_replacement and col_order_match
    # 불필요 인덱스 열: 첫 헤더 동일 + 열 수 동일 + 빈/Unnamed 선두열 없음
    index_ok = (cols_match and (not c_header or c_header[0] == x_header_s[0])
                and not (c_header and (c_header[0] == "" or str(c_header[0]).startswith("Unnamed"))))

    checks = {
        "rows_match": rows_match,
        "cols_match": cols_match,
        "col_order_match": col_order_match,
        "id_sample_match": id_sample_match,
        "hangul_ok": hangul_ok,
        "index_ok": index_ok,
    }
    issues = [k for k, ok in checks.items() if not ok]
    status = "PASS" if not issues else "FAIL"

    return {
        "sheet": ws.title,
        "xlsx_rows": len(x_data),
        "csv_rows": len(c_data),
        "rows_match": rows_match,
        "xlsx_cols": len(x_header_s),
        "csv_cols": len(c_header),
        "cols_match": cols_match,
        "col_order_match": col_order_match,
        "id_column": id_col,
        "id_sample_xlsx": " | ".join(x_ids),
        "id_sample_csv": " | ".join(c_ids),
        "id_sample_match": id_sample_match,
        "hangul_ok": hangul_ok,
        "index_ok": index_ok,
        "status": status,
        "issues": ";".join(issues),
    }


VALIDATION_COLUMNS = [
    "sheet", "xlsx_rows", "csv_rows", "rows_match",
    "xlsx_cols", "csv_cols", "cols_match", "col_order_match",
    "id_column", "id_sample_xlsx", "id_sample_csv", "id_sample_match",
    "hangul_ok", "index_ok", "status", "issues",
]


def run(xlsx_path: Path = XLSX_PATH, out_dir: Path = OUT_DIR,
        validation_path: Path = VALIDATION_PATH) -> list[dict]:
    """전체 변환 + 검증 실행. 검증 결과 리스트 반환."""
    if not xlsx_path.exists():
        raise FileNotFoundError(f"원천 XLSX 없음: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheets = approved_sheets(wb)

    results = []
    for name in sheets:
        ws = wb[name]
        csv_path = convert_sheet(ws, out_dir)
        results.append(validate_sheet(ws, csv_path))

    validation_path.parent.mkdir(parents=True, exist_ok=True)
    with open(validation_path, "w", newline="", encoding=ENCODING) as f:
        w = csv.DictWriter(f, fieldnames=VALIDATION_COLUMNS)
        w.writeheader()
        for r in results:
            w.writerow(r)
    return results


def main() -> int:
    print(f"원천 XLSX: {XLSX_PATH}")
    print(f"출력 폴더: {OUT_DIR}")
    results = run()
    passed = [r for r in results if r["status"] == "PASS"]
    failed = [r for r in results if r["status"] == "FAIL"]

    print(f"\n승인 시트 {len(results)}개 변환 시도 → 성공 {len(passed)} / 실패 {len(failed)}")
    print(f"검증 리포트: {VALIDATION_PATH}")

    for r in passed:
        print(f"  [PASS] {r['sheet']}: {r['csv_rows']}행 x {r['csv_cols']}열")

    if failed:
        print("\n=== 실패 시트 (성공으로 표시하지 않음) ===")
        for r in failed:
            print(f"  [FAIL] {r['sheet']} | 원인: {r['issues']}")
            print(f"         xlsx {r['xlsx_rows']}x{r['xlsx_cols']} vs csv {r['csv_rows']}x{r['csv_cols']}")
        print("\n재실행 방법: 원인 수정 후 `.venv\\Scripts\\python.exe -m src.convert_xlsx` "
              "(원천 XLSX는 수정하지 않음, CSV는 자동 덮어쓰기)")
        return 1

    print("\n모든 승인 시트 검증 통과.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
